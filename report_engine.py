"""
report_engine.py — Основная логика сверочного отчёта для цветочного магазина.
Читает файлы Битрикс (HTML/xls) и Продажи (xlsx), сравнивает, генерирует Excel.
"""

import re
import io
from collections import defaultdict

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# КОНСТАНТЫ
# ──────────────────────────────────────────────────────────────────────────────

MANAGER_MAP = {
    "Аида Акпасова":    "Аида",
    "Томирис Менеджер": "Томирис",
    "Салия Ескараева":  "Салия",
    "Адема Абай":       "Адема",
    "Жулдыз менеджер":  "Жулдыз",
    "Аселя Менеджер":   "Аселя",
    "Менеджер Гульжан": "Гульжан",
    "Улжан Менеджер":   "Улжан",
    "Сая Менеджер":     "Сая",
}

GREEN_STAGES = {"Сделка успешна", "Заказ готов"}

ERROR_ORDER = {
    "Сумма не совпадает":    0,
    "Категории не совпадают": 1,
    "Нет в таблице":         2,
    "Разная дата":           3,
    "Нет в Битрикс":         4,
}

# цвета
C_RED_BG    = "FFC7CE"
C_RED_FG    = "9C0006"
C_YEL_BG    = "FFEB9C"
C_YEL_FG    = "7D4E00"
C_GREEN_BG  = "C6EFCE"
C_GREEN_FG  = "276221"
C_STAGE_GREEN_BG = "EAF3DE"
C_STAGE_RED_BG   = "FFC7CE"
C_HDR_MAIN  = "1F4E79"
C_HDR_MGR   = "2E4057"
C_HDR_COL   = "3D5A80"
C_ERR_TITLE = "7B2226"

FONT_NAME = "Arial"

# ──────────────────────────────────────────────────────────────────────────────
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ──────────────────────────────────────────────────────────────────────────────

def clean_money(val) -> float:
    """Очищает денежное значение и возвращает float."""
    if val is None:
        return 0.0
    s = str(val)
    s = s.replace("₸", "").replace("\u00a0", "").replace(" ", "").replace(",", ".")
    s = s.strip()
    if s in ("", "-", "None", "nan"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_dates_bitrix(raw: str) -> list:
    """Из строки Битрикс извлекает список дат в формате dd.mm.yyyy."""
    if not raw or str(raw).strip() in ("", "None", "nan"):
        return []
    dates = []
    for part in str(raw).split(","):
        part = part.strip()
        # формат dd.mm.yyyy HH:MM:SS
        m = re.match(r"(\d{2}\.\d{2}\.\d{4})", part)
        if m:
            dates.append(m.group(1))
    return dates


def fmt_date(val) -> str:
    """Конвертирует дату в dd.mm.yyyy."""
    if val is None:
        return ""
    s = str(val).strip()
    if s in ("", "None", "nan", "NaT"):
        return ""
    # yyyy-mm-dd
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(3)}.{m.group(2)}.{m.group(1)}"
    # dd.mm.yyyy уже
    m = re.match(r"^(\d{2}\.\d{2}\.\d{4})", s)
    if m:
        return m.group(1)
    return s


def map_manager(bitrix_name: str) -> str:
    """Переводит имя из Битрикс в короткое имя из таблицы Продажи."""
    for k, v in MANAGER_MAP.items():
        if k.lower() == bitrix_name.strip().lower():
            return v
    # попытка частичного совпадения
    for k, v in MANAGER_MAP.items():
        if k.lower() in bitrix_name.strip().lower() or bitrix_name.strip().lower() in k.lower():
            return v
    return bitrix_name.strip()


def stage_color(stage: str):
    """Возвращает (bg, fg) для стадии сделки."""
    if stage in GREEN_STAGES:
        return C_STAGE_GREEN_BG, C_GREEN_FG
    return C_STAGE_RED_BG, C_RED_FG


# ──────────────────────────────────────────────────────────────────────────────
# ЧТЕНИЕ ФАЙЛОВ
# ──────────────────────────────────────────────────────────────────────────────

def read_bitrix(file_bytes: bytes) -> pd.DataFrame:
    """
    Читает HTML-файл Битрикс (Deal_*.xls) через BeautifulSoup.
    Возвращает агрегированный DataFrame (одна строка = одна сделка).
    """
    soup = BeautifulSoup(file_bytes, "lxml")
    table = soup.find("table")
    if table is None:
        raise ValueError("Таблица не найдена в файле Битрикс")

    headers = [th.get_text(strip=True) for th in table.find_all("th")]

    EXPECTED_COLS = [
        "ID", "Стадия сделки", "Ответственный", "Дата и время оплаты",
        "Сумма", "Валюта",
        "Взяли с клиента за букет (сумма без доставки) менеджер 2",
        "Сумма за клубнику менеджер 2",
        "Сумма доставки менеджер 2",
        "Взяли с клиента за букет (сумма без доставки) менеджер 1",
        "Сумма за клубнику менеджер 1",
        "Сумма за доставку менеджер 1",
        "Товар", "Цена", "Количество"
    ]

    rows = []
    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        if not tds:
            continue
        row = [td.get_text(strip=True) for td in tds]
        # дополнить до нужной длины
        while len(row) < len(EXPECTED_COLS):
            row.append("")
        rows.append(row[:len(EXPECTED_COLS)])

    if not rows:
        raise ValueError("Нет данных в файле Битрикс")

    df = pd.DataFrame(rows, columns=EXPECTED_COLS)

    # очистка денег
    money_cols = [
        "Сумма",
        "Взяли с клиента за букет (сумма без доставки) менеджер 2",
        "Сумма за клубнику менеджер 2",
        "Сумма доставки менеджер 2",
        "Взяли с клиента за букет (сумма без доставки) менеджер 1",
        "Сумма за клубнику менеджер 1",
        "Сумма за доставку менеджер 1",
        "Цена", "Количество"
    ]
    for col in money_cols:
        df[col] = df[col].apply(clean_money)

    df["ID"] = df["ID"].astype(str).str.strip()

    # Агрегация по ID
    def agg_bitrix(g):
        raw_date = g["Дата и время оплаты"].iloc[0]
        # если несколько строк — даты могут различаться, собираем уникальные
        all_dates_raw = ", ".join(
            d for d in g["Дата и время оплаты"].unique() if str(d).strip()
        )
        first_date = parse_dates_bitrix(all_dates_raw)
        date_fmt = first_date[0] if first_date else ""

        товары = sorted(set(
            t for t in g["Товар"].unique() if str(t).strip()
        ))

        сумма_товара = (g["Цена"] * g["Количество"]).sum()

        # Шары: если товар содержит "шар" — суммируем Цена*Количество
        шары = (g.loc[g["Товар"].str.lower().str.contains("шар", na=False), "Цена"] *
                g.loc[g["Товар"].str.lower().str.contains("шар", na=False), "Количество"]).sum()

        return pd.Series({
            "Стадия сделки":      g["Стадия сделки"].iloc[0],
            "Ответственный":      g["Ответственный"].iloc[0],
            "Дата и время оплаты": all_dates_raw,
            "Дата_fmt":           date_fmt,
            "Все_даты_б":         set(parse_dates_bitrix(all_dates_raw)),
            "Общая_Битрикс":      g["Сумма"].sum(),
            "М2_цветы":           g["Взяли с клиента за букет (сумма без доставки) менеджер 2"].max(),
            "М2_клубника":        g["Сумма за клубнику менеджер 2"].max(),
            "М2_доставка":        g["Сумма доставки менеджер 2"].max(),
            "М1_цветы":           g["Взяли с клиента за букет (сумма без доставки) менеджер 1"].max(),
            "М1_клубника":        g["Сумма за клубнику менеджер 1"].max(),
            "М1_доставка":        g["Сумма за доставку менеджер 1"].max(),
            "Товары":             ", ".join(товары),
            "Сумма_товара":       сумма_товара,
            "Шары_итого":         шары,
        })

    agg = df.groupby("ID", sort=False).apply(agg_bitrix).reset_index()
    agg["М1_итого"] = agg["М1_цветы"] + agg["М1_клубника"] + agg["М1_доставка"]
    agg["М2_итого"] = agg["М2_цветы"] + agg["М2_клубника"] + agg["М2_доставка"]
    agg["Разница_товар"] = abs(agg["Общая_Битрикс"] - (agg["М1_итого"] + agg["М2_итого"]))
    agg["Менеджер_продажи"] = agg["Ответственный"].apply(map_manager)

    return agg


def read_sales(file_bytes: bytes) -> pd.DataFrame:
    """
    Читает файл Продажи (*.xlsx).
    Возвращает агрегированный DataFrame (одна строка = ID + Менеджер).
    """
    df_raw = pd.read_excel(
        io.BytesIO(file_bytes),
        header=None,
        skiprows=1
    )

    cols = ['Дата', 'ID', 'Менеджер', 'Букет', 'Клубника', 'Парфюм',
            'Доставка', 'Итого', 'Без_дост', 'Оплата', 'Тип', 'Комм', 'extra']

    # обрезаем или дополняем колонки
    if df_raw.shape[1] >= len(cols):
        df_raw = df_raw.iloc[:, :len(cols)]
    else:
        for i in range(len(cols) - df_raw.shape[1]):
            df_raw[df_raw.shape[1] + i] = None

    df_raw.columns = cols

    # фильтр строк-итогов и пустых
    df_raw = df_raw[~df_raw["Менеджер"].isin(["ИТОГО"])]
    df_raw = df_raw[df_raw["Менеджер"].notna()]
    df_raw = df_raw[df_raw["ID"].notna()]

    # конвертация дат
    df_raw["Дата_fmt"] = df_raw["Дата"].apply(fmt_date)

    # очистка денег
    for col in ["Букет", "Клубника", "Доставка", "Итого", "Парфюм"]:
        df_raw[col] = df_raw[col].apply(clean_money)

    df_raw["ID"] = df_raw["ID"].apply(lambda x: str(int(float(x))) if str(x).replace(".", "").isdigit() else str(x)).str.strip()

    # Агрегация по ID + Менеджер
    def agg_sales(g):
        return pd.Series({
            "Букет":     g["Букет"].sum(),
            "Клубника":  g["Клубника"].sum(),
            "Доставка":  g["Доставка"].sum(),
            "Итого":     g["Итого"].sum(),
            "Все_даты":  set(g["Дата_fmt"].dropna().unique()),
            "Дата_fmt":  g["Дата_fmt"].iloc[0],
        })

    agg = df_raw.groupby(["ID", "Менеджер"], sort=False).apply(agg_sales).reset_index()
    return agg


# ──────────────────────────────────────────────────────────────────────────────
# ЛОГИКА СРАВНЕНИЯ
# ──────────────────────────────────────────────────────────────────────────────

def check_dates(dates_b: set, dates_s: set) -> bool:
    """True = даты совпадают (частичные оплаты учтены)."""
    if not dates_b or not dates_s:
        return True  # нет данных — не считаем ошибкой
    return dates_b.issubset(dates_s) or dates_s.issubset(dates_b) or bool(dates_b & dates_s)


def compare(df_b: pd.DataFrame, df_s: pd.DataFrame) -> dict:
    """
    Основное сравнение. Возвращает словарь:
    {
      'errors': [{...}],           # все ошибки для листа Ошибки
      'comparison': {mgr: [{...}]} # данные для листов Сравнение
      'not_in_table': {mgr: [{...}]}
      'not_in_bitrix': {mgr: [{...}]}
    }
    """
    # ID с двумя менеджерами в Продажи
    id_mgr_count = df_s.groupby("ID")["Менеджер"].nunique()
    dual_ids = set(id_mgr_count[id_mgr_count > 1].index)

    bitrix_ids = set(df_b["ID"].unique())
    sales_ids  = set(df_s["ID"].unique())

    errors       = []
    comparison   = defaultdict(list)
    not_in_table = defaultdict(list)
    not_in_bitrix= defaultdict(list)

    # ── проходим по Битрикс ──
    for _, row_b in df_b.iterrows():
        bid   = row_b["ID"]
        stage = row_b["Стадия сделки"]
        resp  = row_b["Ответственный"]
        mgr   = row_b["Менеджер_продажи"]
        total_b = row_b["Общая_Битрикс"]
        dates_b = row_b["Все_даты_б"]
        date_raw= row_b["Дата и время оплаты"]
        date_fmt= row_b["Дата_fmt"]

        цветы_б    = row_b["М1_цветы"] + row_b["М2_цветы"]
        клубника_б = row_b["М1_клубника"] + row_b["М2_клубника"]
        доставка_б = row_b["М1_доставка"] + row_b["М2_доставка"]

        if bid not in sales_ids:
            # Нет в таблице
            errors.append({
                "тип":         "Нет в таблице",
                "стадия":      stage,
                "id":          bid,
                "дата_б":      date_raw,
                "сумма_б":     total_b,
                "сумма_т":     0,
                "расхождение": total_b,
                "менеджер":    mgr,
                "что_делать":  f"Добавить ID {bid} в таблицу продаж. Сумма Битрикс: {total_b:,.0f}₸",
            })
            not_in_table[mgr].append({
                "id": bid, "ответственный": resp,
                "дата": date_fmt,
                "цветы": цветы_б, "клубника": клубника_б,
                "доставка": доставка_б, "шары": row_b["Шары_итого"],
                "итого": total_b,
            })
            # в сравнение тоже добавляем
            comparison[mgr].append(_cmp_row(
                bid, resp, date_fmt, "", dates_b, set(),
                total_b, 0, цветы_б, 0, клубника_б, 0, доставка_б, 0,
                row_b["Шары_итого"], 0, "❌"
            ))
            continue

        # Получаем строки продаж для этого менеджера
        rows_s = df_s[(df_s["ID"] == bid) & (df_s["Менеджер"] == mgr)]

        if bid in dual_ids:
            # Двойной менеджер
            all_rows_s = df_s[df_s["ID"] == bid]
            total_t = all_rows_s["Итого"].sum()
            цветы_т    = all_rows_s["Букет"].sum()
            клубника_т = all_rows_s["Клубника"].sum()
            доставка_т = all_rows_s["Доставка"].sum()
            dates_s    = set().union(*all_rows_s["Все_даты"].tolist())
            date_t     = all_rows_s["Дата_fmt"].iloc[0]

            mgrs_list = list(all_rows_s["Менеджер"].unique())
            mgr_a = mgrs_list[0]
            mgr_b_name = mgrs_list[1] if len(mgrs_list) > 1 else ""
            sum_a = all_rows_s[all_rows_s["Менеджер"] == mgr_a]["Итого"].sum()
            sum_b = all_rows_s[all_rows_s["Менеджер"] == mgr_b_name]["Итого"].sum() if mgr_b_name else 0

            note = f"✅✅ Совпадает (2)\n{mgr_a}: {sum_a:,.0f} / {mgr_b_name}: {sum_b:,.0f} / Итого: {total_t:,.0f}"

            # добавляем в оба листа сравнения
            for m in mgrs_list:
                label = resp if m == mgr else f"{resp} (2-й менеджер)"
                comparison[m].append(_cmp_row(
                    bid, label, date_fmt, date_t, dates_b, dates_s,
                    total_b, total_t, цветы_б, цветы_т,
                    клубника_б, клубника_т, доставка_б, доставка_т,
                    row_b["Шары_итого"], 0, "⚠️", note=note
                ))
            continue

        if rows_s.empty:
            # Есть в Продажи по другому менеджеру — тоже "нет в таблице" для этого
            errors.append({
                "тип":         "Нет в таблице",
                "стадия":      stage,
                "id":          bid,
                "дата_б":      date_raw,
                "сумма_б":     total_b,
                "сумма_т":     0,
                "расхождение": total_b,
                "менеджер":    mgr,
                "что_делать":  f"Добавить ID {bid} в таблицу продаж. Сумма Битрикс: {total_b:,.0f}₸",
            })
            not_in_table[mgr].append({
                "id": bid, "ответственный": resp,
                "дата": date_fmt,
                "цветы": цветы_б, "клубника": клубника_б,
                "доставка": доставка_б, "шары": row_b["Шары_итого"],
                "итого": total_b,
            })
            comparison[mgr].append(_cmp_row(
                bid, resp, date_fmt, "", dates_b, set(),
                total_b, 0, цветы_б, 0, клубника_б, 0, доставка_б, 0,
                row_b["Шары_итого"], 0, "❌"
            ))
            continue

        # Обычное сравнение
        row_s      = rows_s.iloc[0]
        total_t    = row_s["Итого"]
        цветы_т    = row_s["Букет"]
        клубника_т = row_s["Клубника"]
        доставка_т = row_s["Доставка"]
        dates_s    = row_s["Все_даты"]
        date_t     = row_s["Дата_fmt"]

        diff = abs(total_b - total_t)
        date_ok = check_dates(dates_b, dates_s)

        err_type = None
        что_делать = ""

        if diff >= 1:
            err_type = "Сумма не совпадает"
            что_делать = f"Итого Битрикс {total_b:,.0f} → Таблица {total_t:,.0f}"
        elif (abs(цветы_б - цветы_т) >= 1 or
              abs(клубника_б - клубника_т) >= 1 or
              abs(доставка_б - доставка_т) >= 1):
            err_type = "Категории не совпадают"
            parts = []
            if abs(цветы_б - цветы_т) >= 1:
                parts.append(f"Битрикс цветы: {цветы_б:,.0f} → {цветы_т:,.0f}")
            if abs(клубника_б - клубника_т) >= 1:
                parts.append(f"Битрикс клубника: {клубника_б:,.0f} → {клубника_т:,.0f}")
            if abs(доставка_б - доставка_т) >= 1:
                parts.append(f"Битрикс доставка: {доставка_б:,.0f} → {доставка_т:,.0f}")
            что_делать = "; ".join(parts)
        elif not date_ok:
            err_type = "Разная дата"
            б_dates_str = ", ".join(sorted(dates_b))
            т_dates_str = ", ".join(sorted(dates_s))
            что_делать = f"Битрикс: {б_dates_str} | Таблица: {т_dates_str}"

        status = "❌" if err_type else "✅"

        if err_type:
            errors.append({
                "тип":         err_type,
                "стадия":      stage,
                "id":          bid,
                "дата_б":      date_raw,
                "сумма_б":     total_b,
                "сумма_т":     total_t,
                "расхождение": diff,
                "менеджер":    mgr,
                "что_делать":  что_делать,
            })

        comparison[mgr].append(_cmp_row(
            bid, resp, date_fmt, date_t, dates_b, dates_s,
            total_b, total_t, цветы_б, цветы_т,
            клубника_б, клубника_т, доставка_б, доставка_т,
            row_b["Шары_итого"], 0, status
        ))

    # ── проходим по Продажи — ищем "Нет в Битрикс" ──
    for _, row_s in df_s.iterrows():
        sid = row_s["ID"]
        if sid not in bitrix_ids:
            mgr = row_s["Менеджер"]
            total_t = row_s["Итого"]
            errors.append({
                "тип":         "Нет в Битрикс",
                "стадия":      "—",
                "id":          sid,
                "дата_б":      "",
                "сумма_б":     0,
                "сумма_т":     total_t,
                "расхождение": total_t,
                "менеджер":    mgr,
                "что_делать":  f"ID {sid} записан в таблице продаж ({total_t:,.0f}₸), но отсутствует в Битрикс",
            })
            not_in_bitrix[mgr].append({
                "id": sid,
                "дата": row_s["Дата_fmt"],
                "приход": total_t,
                "цветы": row_s["Букет"],
                "клубника": row_s["Клубника"],
                "доставка": row_s["Доставка"],
                "шары": 0,
            })

    return {
        "errors":        errors,
        "comparison":    dict(comparison),
        "not_in_table":  dict(not_in_table),
        "not_in_bitrix": dict(not_in_bitrix),
    }


def _cmp_row(bid, resp, date_b, date_t, dates_b, dates_s,
             total_b, total_t, цветы_б, цветы_т,
             клубника_б, клубника_т, доставка_б, доставка_т,
             шары_б, шары_т, status, note=None):
    date_ok  = check_dates(dates_b, dates_s)
    total_ok = abs(total_b - total_t) < 1
    цв_ok    = abs(цветы_б - цветы_т) < 1
    кл_ok    = abs(клубника_б - клубника_т) < 1
    дост_ok  = abs(доставка_б - доставка_т) < 1
    шары_ok  = abs(шары_б - шары_т) < 1

    def chk(ok, val_b=None, val_t=None):
        if note:
            return note
        if ok:
            return "✅ Совпадает"
        return "❌ Ошибка"

    return {
        "статус":      status,
        "id":          bid,
        "ответственный": resp,
        "дата_б":      date_b,
        "дата_т":      date_t,
        "чк_дата":     chk(date_ok),
        "total_б":     total_b,
        "total_т":     total_t,
        "чк_итого":    chk(total_ok),
        "цветы_б":     цветы_б,
        "цветы_т":     цветы_т,
        "чк_цветы":    chk(цв_ok),
        "клубника_б":  клубника_б,
        "клубника_т":  клубника_т,
        "чк_клубника": chk(кл_ok),
        "доставка_б":  доставка_б,
        "доставка_т":  доставка_т,
        "чк_доставка": chk(дост_ok),
        "шары_б":      шары_б,
        "шары_т":      шары_т,
        "чк_шары":     chk(шары_ok),
        "_date_ok":    date_ok,
        "_total_ok":   total_ok,
        "_цв_ok":      цв_ok,
        "_кл_ok":      кл_ok,
        "_дост_ok":    дост_ok,
        "_шары_ok":    шары_ok,
        "_note":       note,
    }


# ──────────────────────────────────────────────────────────────────────────────
# СТИЛИ OPENPYXL
# ──────────────────────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, name=FONT_NAME):
    return Font(name=name, bold=bold, color=color, size=size)

def align(h="center", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def border_hair():
    s = Side(style="hair", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_thin():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def style_cell(cell, bg=None, fg="000000", bold=False, size=10,
               h_align="center", wrap=True, border=None, num_fmt=None):
    if bg:
        cell.fill = fill(bg)
    cell.font  = font(bold=bold, color=fg, size=size)
    cell.alignment = align(h=h_align, wrap=wrap)
    if border:
        cell.border = border
    if num_fmt:
        cell.number_format = num_fmt


def write_header_row(ws, row_idx, titles, bg, fg="FFFFFF", bold=True,
                     height=30, sizes=None):
    """Записывает строку заголовков колонок."""
    ws.row_dimensions[row_idx].height = height
    for ci, title in enumerate(titles, 1):
        c = ws.cell(row=row_idx, column=ci, value=title)
        style_cell(c, bg=bg, fg=fg, bold=bold, border=border_thin())


def set_col_widths(ws, widths: dict):
    """widths: {col_letter_or_idx: width}"""
    for col, w in widths.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w


# ──────────────────────────────────────────────────────────────────────────────
# ГЕНЕРАЦИЯ EXCEL
# ──────────────────────────────────────────────────────────────────────────────

def generate_excel(df_b: pd.DataFrame, df_s: pd.DataFrame,
                   result: dict) -> bytes:
    wb = Workbook()

    # убираем дефолтный Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _build_errors_sheet(wb, result, df_b)
    _build_product_sheet(wb, df_b)

    managers = sorted(set(
        list(result["comparison"].keys()) +
        list(result["not_in_table"].keys()) +
        list(result["not_in_bitrix"].keys())
    ))

    for mgr in managers:
        _build_comparison_sheet(wb, mgr, result)
        _build_not_in_table_sheet(wb, mgr, result)
        _build_not_in_bitrix_sheet(wb, mgr, result)

    # активный лист — первый
    wb.active = wb.worksheets[0]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Лист "🚨 Ошибки" ──────────────────────────────────────────────────────────

def _build_errors_sheet(wb, result, df_b):
    ws = wb.create_sheet("🚨 Ошибки")
    ws.sheet_view.showGridLines = False

    errors = result["errors"]

    # подсчёт по типам
    counts = defaultdict(int)
    for e in errors:
        counts[e["тип"]] += 1

    # строка 1 — главный заголовок
    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, "🚨 СВЕРОЧНЫЙ ОТЧЁТ — ОШИБКИ")
    style_cell(c, bg=C_ERR_TITLE, fg="FFFFFF", bold=True, size=14)
    ws.row_dimensions[1].height = 40

    # строка 2 — счётчик
    summary = "  |  ".join(
        f"{t}: {counts[t]}" for t in ERROR_ORDER if counts.get(t, 0) > 0
    )
    ws.merge_cells("A2:H2")
    c = ws.cell(2, 1, f"Всего ошибок: {len(errors)}   {summary}")
    style_cell(c, bg="F2F2F2", bold=False, size=10)
    ws.row_dimensions[2].height = 22

    # строка 3 — легенда
    ws.merge_cells("A3:H3")
    c = ws.cell(3, 1,
        "🔴 Красный = Сумма не совпадает / Нет в таблице / Нет в Битрикс   "
        "🟡 Жёлтый = Категории не совпадают / Разная дата")
    style_cell(c, bg="F9F9F9", size=9)
    ws.row_dimensions[3].height = 18

    ws.freeze_panes = "A4"

    COL_TITLES = [
        "Тип ошибки", "Стадия сделки", "ID сделки",
        "Даты в Битрикс", "Сумма Битрикс", "Сумма Таблица",
        "Расхождение ₸", "Что делать"
    ]
    COL_WIDTHS = [22, 18, 12, 28, 16, 16, 16, 50]

    # группируем по менеджерам и сортируем ошибки
    mgr_errors = defaultdict(list)
    for e in errors:
        mgr_errors[e["менеджер"]].append(e)

    cur_row = 4

    for mgr in sorted(mgr_errors.keys()):
        errs = sorted(mgr_errors[mgr],
                      key=lambda x: ERROR_ORDER.get(x["тип"], 99))

        # пустая строка-отступ
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # заголовок менеджера
        ws.merge_cells(f"A{cur_row}:H{cur_row}")
        c = ws.cell(cur_row, 1, mgr)
        style_cell(c, bg=C_HDR_MGR, fg="FFFFFF", bold=True, size=12)
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

        # заголовки колонок
        write_header_row(ws, cur_row, COL_TITLES, C_HDR_COL, height=24)
        cur_row += 1

        for e in errs:
            err_type = e["тип"]
            is_red = err_type in ("Сумма не совпадает", "Нет в таблице", "Нет в Битрикс")
            row_bg = C_RED_BG if is_red else C_YEL_BG
            row_fg = C_RED_FG if is_red else C_YEL_FG

            stage = e["стадия"]
            s_bg, s_fg = stage_color(stage)

            vals = [
                err_type, stage, e["id"], e["дата_б"],
                e["сумма_б"], e["сумма_т"],
                e["расхождение"], e["что_делать"]
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(cur_row, ci, v)
                if ci == 2:  # стадия
                    style_cell(c, bg=s_bg, fg=s_fg, border=border_hair())
                else:
                    style_cell(c, bg=row_bg, fg=row_fg, border=border_hair())
                if ci in (5, 6, 7):
                    c.alignment = align(h="right")
                    if isinstance(v, (int, float)) and v > 0:
                        c.number_format = '#,##0'

            ws.row_dimensions[cur_row].height = 18
            cur_row += 1

    # ширины
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Лист "Товар" ──────────────────────────────────────────────────────────────

def _build_product_sheet(wb, df_b: pd.DataFrame):
    ws = wb.create_sheet("Товар")
    ws.sheet_view.showGridLines = False

    TITLES = [
        "ID", "Стадия сделки", "Ответственный", "Дата и время оплаты",
        "Товары",
        "Цветы(М1)", "Клубника(М1)", "Доставка(М1)",
        "Цветы(М2)", "Клубника(М2)", "Доставка(М2)",
        "Цветы Итого", "Клубника Итого", "Доставка Итого", "Шары Итого",
        "Сумма товара", "Общая сумма Битрикс", "Разница"
    ]
    WIDTHS = [12, 18, 24, 28, 40,
              16, 16, 16, 16, 16, 16,
              16, 16, 16, 16, 16, 18, 12]

    # строка 1 — шапка
    ws.merge_cells(f"A1:{get_column_letter(len(TITLES))}1")
    c = ws.cell(1, 1, "📦 ТОВАРЫ ПО СДЕЛКАМ")
    style_cell(c, bg=C_HDR_MAIN, fg="FFFFFF", bold=True, size=12)
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

    # строка 2 — заголовки
    write_header_row(ws, 2, TITLES, C_HDR_MAIN, height=30)

    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for ri, (_, row) in enumerate(df_b.iterrows(), 3):
        разница = row["Разница_товар"]
        row_bg = C_RED_BG if разница >= 1 else None

        vals = [
            row["ID"],
            row["Стадия сделки"],
            row["Ответственный"],
            row["Дата и время оплаты"],
            row["Товары"],
            row["М1_цветы"], row["М1_клубника"], row["М1_доставка"],
            row["М2_цветы"], row["М2_клубника"], row["М2_доставка"],
            row["М1_цветы"] + row["М2_цветы"],
            row["М1_клубника"] + row["М2_клубника"],
            row["М1_доставка"] + row["М2_доставка"],
            row["Шары_итого"],
            row["Сумма_товара"],
            row["Общая_Битрикс"],
            разница,
        ]

        stage = row["Стадия сделки"]
        s_bg, s_fg = stage_color(stage)

        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            if ci == 2:
                style_cell(c, bg=s_bg, fg=s_fg, border=border_hair())
            elif ci == len(TITLES):  # Разница
                diff_bg = C_RED_BG if разница >= 1 else C_GREEN_BG
                diff_fg = C_RED_FG if разница >= 1 else C_GREEN_FG
                style_cell(c, bg=diff_bg, fg=diff_fg, bold=True,
                           border=border_hair(), num_fmt='#,##0')
            else:
                style_cell(c, bg=row_bg, border=border_hair())
                if ci >= 6 and isinstance(v, (int, float)):
                    c.alignment = align(h="right")
                    c.number_format = '#,##0'

        ws.row_dimensions[ri].height = 18


# ── Лист "Сравнение [Имя]" ────────────────────────────────────────────────────

def _build_comparison_sheet(wb, mgr, result):
    ws = wb.create_sheet(f"Сравнение {mgr}"[:31])
    ws.sheet_view.showGridLines = False

    TITLES = [
        "Статус", "ID", "Ответственный",
        "Дата(Битрикс)", "Дата(Таблица)", "✓Дата",
        "Итого(Б)", "Итого(Т)", "✓Итого",
        "Цветы(Б)", "Цветы(Т)", "✓Цветы",
        "Клубника(Б)", "Клубника(Т)", "✓Клубника",
        "Доставка(Б)", "Доставка(Т)", "✓Доставка",
        "Шары(Б)", "Шары(Т)", "✓Шары"
    ]
    WIDTHS = [8, 12, 24, 18, 18, 22,
              14, 14, 22, 14, 14, 22,
              14, 14, 22, 14, 14, 22,
              14, 14, 22]

    ws.merge_cells(f"A1:{get_column_letter(len(TITLES))}1")
    c = ws.cell(1, 1, f"📋 СРАВНЕНИЕ — {mgr}")
    style_cell(c, bg=C_HDR_MAIN, fg="FFFFFF", bold=True, size=12)
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

    write_header_row(ws, 2, TITLES, C_HDR_MAIN, height=28)

    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    rows = result["comparison"].get(mgr, [])

    for ri, row in enumerate(rows, 3):
        is_dual = row["_note"] is not None
        has_error = row["статус"] == "❌"

        resp_bg = C_RED_BG if has_error else (C_YEL_BG if is_dual else None)
        resp_fg = C_RED_FG if has_error else (C_YEL_FG if is_dual else "000000")

        CHECK_COLS = {
            "чк_дата":     ("_date_ok",  5),
            "чк_итого":    ("_total_ok", 8),
            "чк_цветы":    ("_цв_ok",    11),
            "чк_клубника": ("_кл_ok",    14),
            "чк_доставка": ("_дост_ok",  17),
            "чк_шары":     ("_шары_ok",  20),
        }

        vals = [
            row["статус"], row["id"], row["ответственный"],
            row["дата_б"], row["дата_т"], row["чк_дата"],
            row["total_б"], row["total_т"], row["чк_итого"],
            row["цветы_б"], row["цветы_т"], row["чк_цветы"],
            row["клубника_б"], row["клубника_т"], row["чк_клубника"],
            row["доставка_б"], row["доставка_т"], row["чк_доставка"],
            row["шары_б"], row["шары_т"], row["чк_шары"],
        ]

        chk_positions = {6, 9, 12, 15, 18, 21}  # 1-indexed
        ok_map = {
            6:  row["_date_ok"],
            9:  row["_total_ok"],
            12: row["_цв_ok"],
            15: row["_кл_ok"],
            18: row["_дост_ok"],
            21: row["_шары_ok"],
        }

        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            if ci == 3:  # Ответственный
                style_cell(c, bg=resp_bg, fg=resp_fg,
                           bold=has_error, border=border_hair())
            elif ci in chk_positions:
                ok = ok_map.get(ci, True)
                if is_dual:
                    chk_bg, chk_fg = C_YEL_BG, C_YEL_FG
                elif ok:
                    chk_bg, chk_fg = C_GREEN_BG, C_GREEN_FG
                else:
                    chk_bg, chk_fg = C_RED_BG, C_RED_FG
                style_cell(c, bg=chk_bg, fg=chk_fg, border=border_hair())
            else:
                style_cell(c, border=border_hair())
                if ci in (7, 8, 10, 11, 13, 14, 16, 17, 19, 20):
                    c.alignment = align(h="right")
                    if isinstance(v, (int, float)) and v != 0:
                        c.number_format = '#,##0'

        ws.row_dimensions[ri].height = 18


# ── Лист "Нет в Таблице [Имя]" ───────────────────────────────────────────────

def _build_not_in_table_sheet(wb, mgr, result):
    ws = wb.create_sheet(f"Нет в Таблице {mgr}"[:31])
    ws.sheet_view.showGridLines = False

    TITLES = ["ID", "Ответственный", "Дата", "Цветы",
              "Клубника", "Доставка", "Шары", "Итого"]
    WIDTHS = [12, 24, 18, 16, 16, 16, 16, 16]

    ws.merge_cells(f"A1:{get_column_letter(len(TITLES))}1")
    c = ws.cell(1, 1, f"❌ НЕТ В ТАБЛИЦЕ — {mgr}")
    style_cell(c, bg=C_HDR_MAIN, fg="FFFFFF", bold=True, size=12)
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

    write_header_row(ws, 2, TITLES, C_HDR_MAIN, height=28)

    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    rows = result["not_in_table"].get(mgr, [])
    for ri, row in enumerate(rows, 3):
        vals = [
            row["id"], row["ответственный"], row["дата"],
            row["цветы"], row["клубника"], row["доставка"],
            row["шары"], row["итого"]
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            style_cell(c, bg=C_RED_BG, fg=C_RED_FG, border=border_hair())
            if ci >= 4 and isinstance(v, (int, float)):
                c.alignment = align(h="right")
                c.number_format = '#,##0'
        ws.row_dimensions[ri].height = 18


# ── Лист "Нет в Битрикс [Имя]" ───────────────────────────────────────────────

def _build_not_in_bitrix_sheet(wb, mgr, result):
    ws = wb.create_sheet(f"Нет в Битрикс {mgr}"[:31])
    ws.sheet_view.showGridLines = False

    TITLES = ["ID", "Дата", "Приход", "Цветы",
              "Клубника", "Доставка", "Шары"]
    WIDTHS = [12, 18, 16, 16, 16, 16, 16]

    ws.merge_cells(f"A1:{get_column_letter(len(TITLES))}1")
    c = ws.cell(1, 1, f"❌ НЕТ В БИТРИКС — {mgr}")
    style_cell(c, bg=C_HDR_MAIN, fg="FFFFFF", bold=True, size=12)
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

    write_header_row(ws, 2, TITLES, C_HDR_MAIN, height=28)

    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    rows = result["not_in_bitrix"].get(mgr, [])
    for ri, row in enumerate(rows, 3):
        vals = [
            row["id"], row["дата"], row["приход"],
            row["цветы"], row["клубника"], row["доставка"], row["шары"]
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            style_cell(c, bg=C_RED_BG, fg=C_RED_FG, border=border_hair())
            if ci >= 3 and isinstance(v, (int, float)):
                c.alignment = align(h="right")
                c.number_format = '#,##0'
        ws.row_dimensions[ri].height = 18


# ──────────────────────────────────────────────────────────────────────────────
# ТОЧКА ВХОДА (для запуска напрямую)
# ──────────────────────────────────────────────────────────────────────────────

def run(bitrix_bytes: bytes, sales_bytes: bytes) -> bytes:
    df_b = read_bitrix(bitrix_bytes)
    df_s = read_sales(sales_bytes)
    result = compare(df_b, df_s)
    return generate_excel(df_b, df_s, result)
